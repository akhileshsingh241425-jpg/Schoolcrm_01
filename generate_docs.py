"""
Generate School CRM Documentation:
1. Role Flowchart PDF
2. Roles & Permissions Detail Excel
3. User Manual PDF
"""
import os
from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUTPUT_DIR = os.path.join(os.path.dirname(__file__), 'documentation')
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ======================== DATA ========================

ROLES = {
    'super_admin': {
        'hindi': 'सुपर एडमिन',
        'title': 'Super Administrator',
        'desc': 'CRM platform owner. Manages all schools, subscription plans, billing and platform-level settings.',
        'login': 'Email + Password (school code optional)',
        'modules': ['Super Admin Dashboard', 'Manage Schools', 'Manage Plans', 'Manage Subscriptions', 'All School Modules'],
        'key_tasks': [
            'View platform-wide statistics (total schools, students, staff, revenue)',
            'Add / Edit / Deactivate schools',
            'Create and manage subscription plans (Basic, Standard, Premium)',
            'Assign subscription plans to schools',
            'Toggle school features ON/OFF',
            'Monitor all subscriptions and payment status',
            'Access any school\'s data for support',
        ]
    },
    'school_admin': {
        'hindi': 'स्कूल एडमिन',
        'title': 'School Administrator',
        'desc': 'School-level admin with full access to all modules within their school.',
        'login': 'School Code + Email + Password',
        'modules': ['Dashboard', 'Students', 'Staff', 'Admissions', 'Leads', 'Fees', 'Attendance', 'Academics',
                     'Library', 'Transport', 'Hostel', 'Canteen', 'Sports', 'Health', 'Inventory',
                     'Communication', 'Reports', 'Settings', 'School Branding', 'Data Import'],
        'key_tasks': [
            'Manage all students and staff records',
            'Configure school settings and branding (logo, login background, tagline, theme)',
            'Manage fee structures and track collections',
            'View reports and analytics',
            'Control timetable, exams, and curriculum',
            'Manage all CRM features (leads, admissions, communication)',
            'Import data via Excel/CSV',
            'Assign roles to staff members',
        ]
    },
    'principal': {
        'hindi': 'प्रिंसिपल',
        'title': 'Principal / Vice Principal',
        'desc': 'Senior school authority with read/write access to almost all school modules.',
        'login': 'School Code + Email + Password',
        'modules': ['Dashboard', 'Students', 'Staff', 'Admissions', 'Leads', 'Fees', 'Attendance', 'Academics',
                     'Library', 'Transport', 'Hostel', 'Canteen', 'Sports', 'Health', 'Inventory',
                     'Communication', 'Reports', 'Settings', 'Data Import', 'Parents'],
        'key_tasks': [
            'Oversee all school operations',
            'Review attendance and academic performance',
            'Approve admissions and fee waivers',
            'Monitor staff performance',
            'View all reports',
        ]
    },
    'teacher': {
        'hindi': 'शिक्षक',
        'title': 'Teacher',
        'desc': 'Class teacher or subject teacher with access to academics, attendance and student data.',
        'login': 'School Code + Email + Password',
        'modules': ['Dashboard', 'Students', 'Attendance', 'Academics', 'Library', 'Communication', 'Parents', 'Reports'],
        'key_tasks': [
            'Mark daily attendance for assigned classes',
            'Enter exam marks and grades',
            'View student profiles and parent contacts',
            'Manage timetable and curriculum for own subjects',
            'Send communication to parents',
            'Access library for reference',
        ]
    },
    'accountant': {
        'hindi': 'अकाउंटेंट',
        'title': 'Accountant',
        'desc': 'Handles all financial operations - fee collection, invoices, and financial reports.',
        'login': 'School Code + Email + Password',
        'modules': ['Dashboard', 'Fees', 'Students', 'Inventory', 'Reports'],
        'key_tasks': [
            'Collect fees and issue receipts',
            'Manage fee structures and discounts',
            'Track pending and overdue payments',
            'Generate fee collection reports',
            'Manage school inventory and procurement',
        ]
    },
    'counselor': {
        'hindi': 'काउंसलर',
        'title': 'Counselor / CRM Manager',
        'desc': 'Manages leads, admissions pipeline and parent communication.',
        'login': 'School Code + Email + Password',
        'modules': ['Dashboard', 'Leads', 'Admissions', 'Students', 'Parents', 'Communication'],
        'key_tasks': [
            'Add and follow up on new leads (inquiries)',
            'Convert leads to admissions',
            'Track admission pipeline stages',
            'Communicate with prospective and current parents',
            'Schedule campus visits and counseling sessions',
        ]
    },
    'receptionist': {
        'hindi': 'रिसेप्शनिस्ट',
        'title': 'Receptionist',
        'desc': 'Front desk staff handling visitor inquiries and basic lead/admission entry.',
        'login': 'School Code + Email + Password',
        'modules': ['Dashboard', 'Leads', 'Admissions', 'Students', 'Parents', 'Communication'],
        'key_tasks': [
            'Register new walk-in inquiries as leads',
            'Basic admission form entry',
            'Provide student and parent information',
            'Send basic communication/notifications',
        ]
    },
    'exam_controller': {
        'hindi': 'परीक्षा नियंत्रक',
        'title': 'Exam Controller / Coordinator',
        'desc': 'Manages exam schedules, results, and report card generation.',
        'login': 'School Code + Email + Password',
        'modules': ['Dashboard', 'Academics', 'Students', 'Reports'],
        'key_tasks': [
            'Create and manage exam schedules',
            'Enter and verify exam results',
            'Generate report cards',
            'Manage grading system and grade mapping',
        ]
    },
    'department_head': {
        'hindi': 'विभागाध्यक्ष (HOD)',
        'title': 'Department Head / HOD',
        'desc': 'Head of Department with access to academics and attendance for their department.',
        'login': 'School Code + Email + Password',
        'modules': ['Dashboard', 'Academics', 'Attendance', 'Students', 'Library', 'Communication', 'Reports'],
        'key_tasks': [
            'Oversee department curriculum',
            'Review attendance of department teachers',
            'Monitor student performance in department subjects',
            'Coordinate with library for book requirements',
        ]
    },
    'hr_manager': {
        'hindi': 'HR मैनेजर',
        'title': 'HR Manager',
        'desc': 'Manages staff records, attendance, leave, payroll and HR operations.',
        'login': 'School Code + Email + Password',
        'modules': ['Dashboard', 'Staff', 'Attendance', 'Communication', 'Reports'],
        'key_tasks': [
            'Manage staff profiles and documents',
            'Track staff attendance and leave',
            'Process payroll and salary slips',
            'Handle staff recruitment and onboarding',
            'Generate HR reports',
        ]
    },
    'hostel_warden': {
        'hindi': 'हॉस्टल वार्डन',
        'title': 'Hostel Warden',
        'desc': 'Manages hostel rooms, student allocation, and hostel attendance.',
        'login': 'School Code + Email + Password',
        'modules': ['Dashboard', 'Hostel', 'Students', 'Attendance', 'Health', 'Communication'],
        'key_tasks': [
            'Manage hostel rooms and bed allocation',
            'Track hostel student attendance',
            'Handle hostel complaints and maintenance',
            'Monitor student health in hostel',
            'Communicate with parents about hostel matters',
        ]
    },
    'transport_manager': {
        'hindi': 'ट्रांसपोर्ट मैनेजर',
        'title': 'Transport Manager',
        'desc': 'Manages school buses, routes, drivers, and student transport allocation.',
        'login': 'School Code + Email + Password',
        'modules': ['Dashboard', 'Transport', 'Students', 'Communication', 'Reports'],
        'key_tasks': [
            'Manage bus routes and stops',
            'Assign students to transport routes',
            'Manage drivers and vehicle records',
            'Track transport fee collection',
            'Generate transport reports',
        ]
    },
    'librarian': {
        'hindi': 'लाइब्रेरियन',
        'title': 'Librarian',
        'desc': 'Manages library books, issue/return, and library reports.',
        'login': 'School Code + Email + Password',
        'modules': ['Dashboard', 'Library', 'Students', 'Reports'],
        'key_tasks': [
            'Add and manage book records',
            'Issue and return books',
            'Track overdue books and fines',
            'Generate library usage reports',
        ]
    },
    'canteen_manager': {
        'hindi': 'कैंटीन मैनेजर',
        'title': 'Canteen Manager',
        'desc': 'Manages canteen menu, orders, and canteen inventory.',
        'login': 'School Code + Email + Password',
        'modules': ['Dashboard', 'Canteen', 'Inventory', 'Reports'],
        'key_tasks': [
            'Manage daily menu items',
            'Process canteen orders',
            'Track canteen inventory and expenses',
            'Generate canteen sales reports',
        ]
    },
    'sports_incharge': {
        'hindi': 'खेल प्रभारी (PTI)',
        'title': 'Sports In-Charge / PTI',
        'desc': 'Manages sports activities, teams, and achievements.',
        'login': 'School Code + Email + Password',
        'modules': ['Dashboard', 'Sports', 'Students', 'Health', 'Communication'],
        'key_tasks': [
            'Manage sports events and teams',
            'Record student sports achievements',
            'Track student fitness and health records',
            'Schedule practice sessions',
        ]
    },
    'health_officer': {
        'hindi': 'स्वास्थ्य अधिकारी',
        'title': 'Health Officer / Nurse',
        'desc': 'Manages student health records, medical visits, and health reports.',
        'login': 'School Code + Email + Password',
        'modules': ['Dashboard', 'Health', 'Students', 'Communication', 'Reports'],
        'key_tasks': [
            'Record student health checkup data',
            'Track medical visits and prescriptions',
            'Manage vaccination records',
            'Generate health reports',
            'Communicate with parents about health issues',
        ]
    },
    'lab_assistant': {
        'hindi': 'लैब असिस्टेंट',
        'title': 'Lab Assistant',
        'desc': 'Manages lab equipment inventory and assists in academic labs.',
        'login': 'School Code + Email + Password',
        'modules': ['Dashboard', 'Academics', 'Inventory'],
        'key_tasks': [
            'Track lab equipment and chemicals inventory',
            'Assist in practical exams',
            'Maintain lab usage records',
        ]
    },
    'parent': {
        'hindi': 'अभिभावक',
        'title': 'Parent',
        'desc': 'Parent portal to view child\'s information and communicate with school.',
        'login': 'School Code + Email/Phone + Password',
        'modules': ['Dashboard (My Children)', 'Communication'],
        'key_tasks': [
            'View child\'s attendance and academics',
            'Check fee payment status',
            'Receive school notifications',
            'Communicate with teachers',
            'View exam results and report cards',
        ]
    },
    'student': {
        'hindi': 'छात्र',
        'title': 'Student',
        'desc': 'Student self-service portal (limited access).',
        'login': 'School Code + Email + Password',
        'modules': ['Dashboard'],
        'key_tasks': [
            'View own attendance',
            'View exam results',
            'Access timetable',
            'View library books issued',
        ]
    },
}

ALL_MODULES = [
    ('Dashboard', 'Main dashboard with overview statistics and quick links'),
    ('Students', 'Student registration, profiles, documents, class-wise listing'),
    ('Staff', 'Staff profiles, documents, departments, designations'),
    ('Admissions', 'Admission pipeline - application to enrollment'),
    ('Leads', 'CRM leads management - inquiries, follow-ups, conversion'),
    ('Fees', 'Fee structures, collection, receipts, dues tracking'),
    ('Attendance', 'Student & staff daily attendance marking and reports'),
    ('Academics', 'Timetable, curriculum, exams, results, report cards'),
    ('Library', 'Book management, issue/return, overdue tracking'),
    ('Transport', 'Bus routes, stops, student allocation, vehicle tracking'),
    ('Hostel', 'Room management, student allocation, hostel attendance'),
    ('Canteen', 'Menu management, orders, canteen inventory'),
    ('Sports', 'Sports events, teams, achievements, fitness records'),
    ('Health', 'Student health records, medical visits, vaccinations'),
    ('Inventory', 'School inventory, procurement, stock tracking'),
    ('Communication', 'Notices, SMS, email, WhatsApp messaging'),
    ('Reports', 'Admission, fee, attendance, marketing analytics'),
    ('Settings', 'School settings, user management, role configuration'),
    ('School Branding', 'Logo, login background, banner, tagline, theme customization'),
    ('Data Import', 'Bulk import students/staff via Excel/CSV'),
    ('Parents', 'Parent profiles and communication'),
    ('Super Admin Portal', 'Platform-level management (Super Admin only)'),
]


# ======================== 1. FLOWCHART PDF ========================

def create_flowchart_pdf():
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    
    # Page 1: Title
    pdf.add_page('L')
    pdf.set_font('Helvetica', 'B', 28)
    pdf.cell(0, 20, '', ln=True)
    pdf.cell(0, 15, 'School CRM - Role Hierarchy & Flowchart', align='C', ln=True)
    pdf.set_font('Helvetica', '', 14)
    pdf.cell(0, 10, 'Complete Role-Based Access Control (RBAC) Structure', align='C', ln=True)
    pdf.cell(0, 8, '', ln=True)
    
    # Draw hierarchy
    pdf.set_font('Helvetica', 'B', 11)
    
    # Level 0: Super Admin
    pdf.set_fill_color(220, 53, 69)  # Red
    pdf.set_text_color(255, 255, 255)
    x_center = 140
    pdf.set_xy(x_center - 35, 65)
    pdf.cell(70, 12, 'SUPER ADMIN', border=1, align='C', fill=True)
    
    # Arrow down
    pdf.set_text_color(0, 0, 0)
    pdf.set_xy(x_center - 1, 77)
    pdf.cell(2, 8, '|', align='C')
    pdf.set_xy(x_center - 1, 83)
    pdf.cell(2, 5, 'V', align='C')
    
    # Level 1: School Admin
    pdf.set_fill_color(0, 123, 255)  # Blue
    pdf.set_text_color(255, 255, 255)
    pdf.set_xy(x_center - 35, 90)
    pdf.cell(70, 12, 'SCHOOL ADMIN', border=1, align='C', fill=True)
    
    # Arrow down
    pdf.set_text_color(0, 0, 0)
    pdf.set_xy(x_center - 1, 102)
    pdf.cell(2, 8, '|', align='C')
    
    # Level 2: Management roles
    mgmt_roles = ['Principal', 'HR Manager']
    pdf.set_fill_color(40, 167, 69)  # Green
    pdf.set_text_color(255, 255, 255)
    start_x = x_center - 55
    for i, role in enumerate(mgmt_roles):
        pdf.set_xy(start_x + i * 60, 115)
        pdf.cell(50, 10, role, border=1, align='C', fill=True)
    
    # Arrow down
    pdf.set_text_color(0, 0, 0)
    pdf.set_xy(x_center - 1, 125)
    pdf.cell(2, 8, '|', align='C')
    
    # Level 3: Academic roles
    acad_roles = ['Teacher', 'Exam Controller', 'HOD', 'Librarian']
    pdf.set_fill_color(255, 193, 7)  # Yellow
    pdf.set_text_color(0, 0, 0)
    start_x = x_center - 100
    for i, role in enumerate(acad_roles):
        pdf.set_xy(start_x + i * 52, 138)
        pdf.cell(48, 10, role, border=1, align='C', fill=True)
    
    # Level 4: Operational roles
    pdf.set_xy(x_center - 1, 148)
    pdf.cell(2, 6, '|', align='C')
    
    ops_roles = ['Accountant', 'Counselor', 'Receptionist', 'Transport Mgr', 'Hostel Warden']
    pdf.set_fill_color(23, 162, 184)  # Cyan
    pdf.set_text_color(255, 255, 255)
    start_x = x_center - 130
    for i, role in enumerate(ops_roles):
        pdf.set_xy(start_x + i * 54, 158)
        pdf.cell(50, 10, role, border=1, align='C', fill=True)
    
    # Level 5: Specialist roles
    pdf.set_text_color(0, 0, 0)
    pdf.set_xy(x_center - 1, 168)
    pdf.cell(2, 6, '|', align='C')
    
    spec_roles = ['Canteen Mgr', 'Sports PTI', 'Health Officer', 'Lab Assistant']
    pdf.set_fill_color(108, 117, 125)  # Gray
    pdf.set_text_color(255, 255, 255)
    start_x = x_center - 100
    for i, role in enumerate(spec_roles):
        pdf.set_xy(start_x + i * 52, 178)
        pdf.cell(48, 10, role, border=1, align='C', fill=True)
    
    # Level 6: End users
    pdf.set_text_color(0, 0, 0)
    pdf.set_xy(x_center - 1, 188)
    pdf.cell(2, 6, '|', align='C')
    
    end_roles = ['Parent', 'Student']
    pdf.set_fill_color(111, 66, 193)  # Purple
    pdf.set_text_color(255, 255, 255)
    start_x = x_center - 35
    for i, role in enumerate(end_roles):
        pdf.set_xy(start_x + i * 40, 198)
        pdf.cell(35, 10, role, border=1, align='C', fill=True)
    
    pdf.set_text_color(0, 0, 0)
    
    # Legend
    pdf.set_font('Helvetica', 'B', 10)
    pdf.set_xy(10, 75)
    pdf.cell(0, 6, 'LEGEND:', ln=True)
    pdf.set_font('Helvetica', '', 9)
    
    legend = [
        ((220, 53, 69), 'Platform Level (CRM Owner)'),
        ((0, 123, 255), 'School Level (Full Access)'),
        ((40, 167, 69), 'Management Level'),
        ((255, 193, 7), 'Academic Level'),
        ((23, 162, 184), 'Operations Level'),
        ((108, 117, 125), 'Specialist Level'),
        ((111, 66, 193), 'End User Level'),
    ]
    y = 82
    for color, label in legend:
        pdf.set_fill_color(*color)
        pdf.set_xy(12, y)
        pdf.cell(8, 5, '', border=1, fill=True)
        pdf.set_xy(22, y)
        pdf.cell(50, 5, label)
        y += 7
    
    # Page 2: Login Flow
    pdf.add_page('L')
    pdf.set_font('Helvetica', 'B', 20)
    pdf.cell(0, 12, 'Login Flow', align='C', ln=True)
    pdf.cell(0, 5, '', ln=True)
    
    pdf.set_font('Helvetica', 'B', 11)
    pdf.set_fill_color(240, 240, 240)
    pdf.cell(50, 8, 'Step', border=1, fill=True, align='C')
    pdf.cell(100, 8, 'Super Admin Login', border=1, fill=True, align='C')
    pdf.cell(120, 8, 'School User Login (Admin/Teacher/Staff etc.)', border=1, fill=True, align='C')
    pdf.ln()
    
    pdf.set_font('Helvetica', '', 10)
    steps = [
        ('Step 1', 'Open login page', 'Open login page'),
        ('Step 2', 'Enter Email: admin@schoolcrm.com', 'Enter School Code (e.g. Sikarwar)'),
        ('Step 3', 'Enter Password: superadmin123', 'Enter Email & Password'),
        ('Step 4', 'School Code: Any or Blank', 'Click Login'),
        ('Step 5', 'Click Login', 'Redirected to School Dashboard'),
        ('Step 6', 'Redirected to Super Admin Portal', 'Access modules based on role'),
    ]
    for step, sa, su in steps:
        pdf.cell(50, 8, step, border=1, align='C')
        pdf.cell(100, 8, sa, border=1, align='C')
        pdf.cell(120, 8, su, border=1, align='C')
        pdf.ln()
    
    # Page 3: Role -> Module Access Matrix
    pdf.add_page('L')
    pdf.set_font('Helvetica', 'B', 18)
    pdf.cell(0, 12, 'Role - Module Access Matrix', align='C', ln=True)
    pdf.cell(0, 3, '', ln=True)
    
    modules_short = ['Dashboard', 'Students', 'Staff', 'Admissions', 'Leads', 'Fees', 'Attendance', 
                     'Academics', 'Library', 'Transport', 'Hostel', 'Canteen', 'Sports', 'Health', 
                     'Inventory', 'Comm.', 'Reports', 'Settings', 'Import', 'Parents']
    
    role_module_map = {
        'Super Admin': ['Dashboard', 'Students', 'Staff', 'Admissions', 'Leads', 'Fees', 'Attendance', 'Academics', 'Library', 'Transport', 'Hostel', 'Canteen', 'Sports', 'Health', 'Inventory', 'Comm.', 'Reports', 'Settings', 'Import', 'Parents'],
        'School Admin': ['Dashboard', 'Students', 'Staff', 'Admissions', 'Leads', 'Fees', 'Attendance', 'Academics', 'Library', 'Transport', 'Hostel', 'Canteen', 'Sports', 'Health', 'Inventory', 'Comm.', 'Reports', 'Settings', 'Import', 'Parents'],
        'Principal': ['Dashboard', 'Students', 'Staff', 'Admissions', 'Leads', 'Fees', 'Attendance', 'Academics', 'Library', 'Transport', 'Hostel', 'Canteen', 'Sports', 'Health', 'Inventory', 'Comm.', 'Reports', 'Settings', 'Import', 'Parents'],
        'Teacher': ['Dashboard', 'Students', '', '', '', '', 'Attendance', 'Academics', 'Library', '', '', '', '', '', '', 'Comm.', 'Reports', '', '', 'Parents'],
        'Accountant': ['Dashboard', 'Students', '', '', '', 'Fees', '', '', '', '', '', '', '', '', 'Inventory', '', 'Reports', '', '', ''],
        'Counselor': ['Dashboard', 'Students', '', 'Admissions', 'Leads', '', '', '', '', '', '', '', '', '', '', 'Comm.', '', '', '', 'Parents'],
        'Receptionist': ['Dashboard', 'Students', '', 'Admissions', 'Leads', '', '', '', '', '', '', '', '', '', '', 'Comm.', '', '', '', 'Parents'],
        'Exam Controller': ['Dashboard', 'Students', '', '', '', '', '', 'Academics', '', '', '', '', '', '', '', '', 'Reports', '', '', ''],
        'HOD': ['Dashboard', 'Students', '', '', '', '', 'Attendance', 'Academics', 'Library', '', '', '', '', '', '', 'Comm.', 'Reports', '', '', ''],
        'HR Manager': ['Dashboard', '', 'Staff', '', '', '', 'Attendance', '', '', '', '', '', '', '', '', 'Comm.', 'Reports', '', '', ''],
        'Hostel Warden': ['Dashboard', 'Students', '', '', '', '', 'Attendance', '', '', '', 'Hostel', '', '', 'Health', '', 'Comm.', '', '', '', ''],
        'Transport Mgr': ['Dashboard', 'Students', '', '', '', '', '', '', '', 'Transport', '', '', '', '', '', 'Comm.', 'Reports', '', '', ''],
        'Librarian': ['Dashboard', 'Students', '', '', '', '', '', '', 'Library', '', '', '', '', '', '', '', 'Reports', '', '', ''],
        'Canteen Mgr': ['Dashboard', '', '', '', '', '', '', '', '', '', '', 'Canteen', '', '', 'Inventory', '', 'Reports', '', '', ''],
        'Sports PTI': ['Dashboard', 'Students', '', '', '', '', '', '', '', '', '', '', 'Sports', 'Health', '', 'Comm.', '', '', '', ''],
        'Health Officer': ['Dashboard', 'Students', '', '', '', '', '', '', '', '', '', '', '', 'Health', '', 'Comm.', 'Reports', '', '', ''],
        'Lab Assistant': ['Dashboard', '', '', '', '', '', '', 'Academics', '', '', '', '', '', '', 'Inventory', '', '', '', '', ''],
        'Parent': ['Dashboard', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'Comm.', '', '', '', ''],
        'Student': ['Dashboard', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', ''],
    }
    
    # Header
    pdf.set_font('Helvetica', 'B', 6)
    pdf.set_fill_color(44, 62, 80)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(28, 7, 'Role', border=1, fill=True, align='C')
    col_w = 12.1
    for m in modules_short:
        pdf.cell(col_w, 7, m, border=1, fill=True, align='C')
    pdf.ln()
    
    # Rows
    pdf.set_font('Helvetica', '', 6)
    pdf.set_text_color(0, 0, 0)
    row_idx = 0
    for role, access in role_module_map.items():
        if row_idx % 2 == 0:
            pdf.set_fill_color(245, 245, 245)
        else:
            pdf.set_fill_color(255, 255, 255)
        
        pdf.set_font('Helvetica', 'B', 6)
        pdf.cell(28, 6, role, border=1, fill=True)
        pdf.set_font('Helvetica', '', 6)
        for i, m in enumerate(modules_short):
            if m and access[i]:
                pdf.set_fill_color(200, 230, 200)
                pdf.cell(col_w, 6, 'Yes', border=1, fill=True, align='C')
            else:
                pdf.set_fill_color(255, 220, 220)
                pdf.cell(col_w, 6, '-', border=1, fill=True, align='C')
            if row_idx % 2 == 0:
                pdf.set_fill_color(245, 245, 245)
            else:
                pdf.set_fill_color(255, 255, 255)
        pdf.ln()
        row_idx += 1
    
    # Page 4: Each role detail
    pdf.add_page('P')
    pdf.set_font('Helvetica', 'B', 20)
    pdf.cell(0, 12, 'Role Details', align='C', ln=True)
    pdf.cell(0, 5, '', ln=True)
    
    for role_key, info in ROLES.items():
        if pdf.get_y() > 240:
            pdf.add_page('P')
        
        pdf.set_font('Helvetica', 'B', 12)
        pdf.set_fill_color(44, 62, 80)
        pdf.set_text_color(255, 255, 255)
        pdf.cell(0, 8, f"  {info['title']} ({role_key})", border=1, fill=True, ln=True)
        pdf.set_text_color(0, 0, 0)
        
        pdf.set_font('Helvetica', '', 9)
        pdf.cell(0, 6, f"Description: {info['desc']}", ln=True)
        pdf.cell(0, 6, f"Login Method: {info['login']}", ln=True)
        pdf.cell(0, 6, f"Modules: {', '.join(info['modules'])}", ln=True)
        pdf.cell(0, 5, 'Key Tasks:', ln=True)
        for task in info['key_tasks']:
            pdf.cell(8)
            pdf.cell(0, 5, f"- {task}", ln=True)
        pdf.cell(0, 4, '', ln=True)
    
    filepath = os.path.join(OUTPUT_DIR, 'School_CRM_Role_Flowchart.pdf')
    pdf.output(filepath)
    print(f'[OK] Flowchart PDF: {filepath}')


# ======================== 2. EXCEL ========================

def create_roles_excel():
    wb = Workbook()
    
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    wrap_align = Alignment(wrap_text=True, vertical='top')
    
    # Sheet 1: Roles Overview
    ws1 = wb.active
    ws1.title = 'Roles Overview'
    headers = ['#', 'Role Key', 'Role Title', 'Description', 'Login Method', 'Accessible Modules', 'Key Responsibilities']
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    ws1.column_dimensions['A'].width = 5
    ws1.column_dimensions['B'].width = 18
    ws1.column_dimensions['C'].width = 25
    ws1.column_dimensions['D'].width = 50
    ws1.column_dimensions['E'].width = 30
    ws1.column_dimensions['F'].width = 60
    ws1.column_dimensions['G'].width = 60
    
    for i, (role_key, info) in enumerate(ROLES.items(), 1):
        row = i + 1
        ws1.cell(row=row, column=1, value=i).border = thin_border
        ws1.cell(row=row, column=2, value=role_key).border = thin_border
        ws1.cell(row=row, column=3, value=info['title']).border = thin_border
        ws1.cell(row=row, column=4, value=info['desc']).border = thin_border
        ws1.cell(row=row, column=4).alignment = wrap_align
        ws1.cell(row=row, column=5, value=info['login']).border = thin_border
        ws1.cell(row=row, column=6, value=', '.join(info['modules'])).border = thin_border
        ws1.cell(row=row, column=6).alignment = wrap_align
        ws1.cell(row=row, column=7, value='\n'.join(f"- {t}" for t in info['key_tasks'])).border = thin_border
        ws1.cell(row=row, column=7).alignment = wrap_align
        ws1.row_dimensions[row].height = max(30, len(info['key_tasks']) * 15)
        
        # Color code
        if role_key == 'super_admin':
            fill = PatternFill(start_color='FADBD8', end_color='FADBD8', fill_type='solid')
        elif role_key == 'school_admin':
            fill = PatternFill(start_color='D6EAF8', end_color='D6EAF8', fill_type='solid')
        elif role_key in ('principal', 'hr_manager'):
            fill = PatternFill(start_color='D5F5E3', end_color='D5F5E3', fill_type='solid')
        elif role_key in ('parent', 'student'):
            fill = PatternFill(start_color='E8DAEF', end_color='E8DAEF', fill_type='solid')
        else:
            fill = PatternFill(start_color='FDEBD0', end_color='FDEBD0', fill_type='solid')
        
        for col in range(1, 8):
            ws1.cell(row=row, column=col).fill = fill
    
    # Sheet 2: Module Access Matrix
    ws2 = wb.create_sheet('Module Access Matrix')
    modules_list = ['Dashboard', 'Students', 'Staff', 'Admissions', 'Leads', 'Fees', 'Attendance',
                    'Academics', 'Library', 'Transport', 'Hostel', 'Canteen', 'Sports', 'Health',
                    'Inventory', 'Communication', 'Reports', 'Settings', 'Data Import', 'Parents']
    
    role_db_modules = {
        'super_admin': ['dashboard', 'students', 'staff', 'admissions', 'leads', 'fees', 'attendance', 'academics', 'library', 'transport', 'hostel', 'canteen', 'sports', 'health', 'inventory', 'communication', 'reports', 'settings', 'data_import', 'parents'],
        'school_admin': ['dashboard', 'students', 'staff', 'admissions', 'leads', 'fees', 'attendance', 'academics', 'library', 'transport', 'hostel', 'canteen', 'sports', 'health', 'inventory', 'communication', 'reports', 'settings', 'data_import', 'parents'],
        'principal': ['dashboard', 'students', 'staff', 'admissions', 'leads', 'fees', 'attendance', 'academics', 'library', 'transport', 'hostel', 'canteen', 'sports', 'health', 'inventory', 'communication', 'reports', 'settings', 'data_import', 'parents'],
        'teacher': ['dashboard', 'students', '', '', '', '', 'attendance', 'academics', 'library', '', '', '', '', '', '', 'communication', 'reports', '', '', 'parents'],
        'accountant': ['dashboard', 'students', '', '', '', 'fees', '', '', '', '', '', '', '', '', 'inventory', '', 'reports', '', '', ''],
        'counselor': ['dashboard', 'students', '', 'admissions', 'leads', '', '', '', '', '', '', '', '', '', '', 'communication', '', '', '', 'parents'],
        'receptionist': ['dashboard', 'students', '', 'admissions', 'leads', '', '', '', '', '', '', '', '', '', '', 'communication', '', '', '', 'parents'],
        'exam_controller': ['dashboard', 'students', '', '', '', '', '', 'academics', '', '', '', '', '', '', '', '', 'reports', '', '', ''],
        'department_head': ['dashboard', 'students', '', '', '', '', 'attendance', 'academics', 'library', '', '', '', '', '', '', 'communication', 'reports', '', '', ''],
        'hr_manager': ['dashboard', '', 'staff', '', '', '', 'attendance', '', '', '', '', '', '', '', '', 'communication', 'reports', '', '', ''],
        'hostel_warden': ['dashboard', 'students', '', '', '', '', 'attendance', '', '', '', 'hostel', '', '', 'health', '', 'communication', '', '', '', ''],
        'transport_manager': ['dashboard', 'students', '', '', '', '', '', '', '', 'transport', '', '', '', '', '', 'communication', 'reports', '', '', ''],
        'librarian': ['dashboard', 'students', '', '', '', '', '', '', 'library', '', '', '', '', '', '', '', 'reports', '', '', ''],
        'canteen_manager': ['dashboard', '', '', '', '', '', '', '', '', '', '', 'canteen', '', '', 'inventory', '', 'reports', '', '', ''],
        'sports_incharge': ['dashboard', 'students', '', '', '', '', '', '', '', '', '', '', 'sports', 'health', '', 'communication', '', '', '', ''],
        'health_officer': ['dashboard', 'students', '', '', '', '', '', '', '', '', '', '', '', 'health', '', 'communication', 'reports', '', '', ''],
        'lab_assistant': ['dashboard', '', '', '', '', '', '', 'academics', '', '', '', '', '', '', 'inventory', '', '', '', '', ''],
        'parent': ['dashboard', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'communication', '', '', '', ''],
        'student': ['dashboard', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', ''],
    }
    
    # Header
    ws2.cell(row=1, column=1, value='Role').font = header_font
    ws2.cell(row=1, column=1).fill = header_fill
    ws2.cell(row=1, column=1).border = thin_border
    ws2.column_dimensions['A'].width = 20
    
    for col, mod in enumerate(modules_list, 2):
        cell = ws2.cell(row=1, column=col, value=mod)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(text_rotation=45, horizontal='center')
        ws2.column_dimensions[cell.column_letter].width = 12
    
    green_fill = PatternFill(start_color='27AE60', end_color='27AE60', fill_type='solid')
    red_fill = PatternFill(start_color='E74C3C', end_color='E74C3C', fill_type='solid')
    
    for i, (role_key, access) in enumerate(role_db_modules.items(), 1):
        row = i + 1
        ws2.cell(row=row, column=1, value=role_key).border = thin_border
        ws2.cell(row=row, column=1).font = Font(bold=True)
        
        for col, mod_access in enumerate(access, 2):
            cell = ws2.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
            if mod_access:
                cell.value = 'YES'
                cell.fill = green_fill
                cell.font = Font(color='FFFFFF', bold=True, size=8)
            else:
                cell.value = '-'
                cell.fill = red_fill
                cell.font = Font(color='FFFFFF', size=8)
    
    # Sheet 3: All Modules Description
    ws3 = wb.create_sheet('Modules Description')
    headers3 = ['#', 'Module Name', 'Description', 'Accessible By Roles']
    for col, h in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    ws3.column_dimensions['A'].width = 5
    ws3.column_dimensions['B'].width = 20
    ws3.column_dimensions['C'].width = 50
    ws3.column_dimensions['D'].width = 70
    
    mod_name_lower_map = {m.lower().replace(' ', '_'): m for m, _ in ALL_MODULES}
    
    for i, (mod_name, mod_desc) in enumerate(ALL_MODULES, 1):
        row = i + 1
        ws3.cell(row=row, column=1, value=i).border = thin_border
        ws3.cell(row=row, column=2, value=mod_name).border = thin_border
        ws3.cell(row=row, column=2).font = Font(bold=True)
        ws3.cell(row=row, column=3, value=mod_desc).border = thin_border
        ws3.cell(row=row, column=3).alignment = wrap_align
        
        # Find roles that have access
        mod_lower = mod_name.lower().replace(' ', '_')
        accessible_roles = []
        for rk, info in ROLES.items():
            for m in info['modules']:
                if mod_name.lower() in m.lower() or m.lower().replace(' ', '_') == mod_lower:
                    accessible_roles.append(info['title'])
                    break
        ws3.cell(row=row, column=4, value=', '.join(accessible_roles) if accessible_roles else 'Super Admin Only').border = thin_border
        ws3.cell(row=row, column=4).alignment = wrap_align
    
    # Sheet 4: Login Credentials
    ws4 = wb.create_sheet('Login Credentials')
    headers4 = ['Role', 'School Code', 'Email', 'Password', 'Redirect Page']
    for col, h in enumerate(headers4, 1):
        cell = ws4.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    ws4.column_dimensions['A'].width = 20
    ws4.column_dimensions['B'].width = 15
    ws4.column_dimensions['C'].width = 35
    ws4.column_dimensions['D'].width = 18
    ws4.column_dimensions['E'].width = 25
    
    creds = [
        ['Super Admin', 'Any / Blank', 'admin@schoolcrm.com', 'superadmin123', '/super-admin'],
        ['School Admin', 'Sikarwar', 'akhileshsingh241425@gmail.com', 'admin123', '/dashboard'],
        ['Parent', 'Sikarwar', '(as created)', '(as set)', '/my-children'],
        ['Teacher/Staff', 'Sikarwar', '(as created)', '(as set)', '/dashboard'],
    ]
    for i, row_data in enumerate(creds, 2):
        for col, val in enumerate(row_data, 1):
            cell = ws4.cell(row=i, column=col, value=val)
            cell.border = thin_border
    
    filepath = os.path.join(OUTPUT_DIR, 'School_CRM_Roles_Details.xlsx')
    wb.save(filepath)
    print(f'[OK] Excel: {filepath}')


# ======================== 3. USER MANUAL PDF ========================

def create_user_manual():
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    
    def section_title(text):
        pdf.set_font('Helvetica', 'B', 16)
        pdf.set_fill_color(44, 62, 80)
        pdf.set_text_color(255, 255, 255)
        pdf.cell(0, 10, f'  {text}', fill=True, ln=True)
        pdf.set_text_color(0, 0, 0)
        pdf.cell(0, 3, '', ln=True)
    
    def sub_title(text):
        pdf.set_font('Helvetica', 'B', 12)
        pdf.set_text_color(44, 62, 80)
        pdf.cell(0, 8, text, ln=True)
        pdf.set_text_color(0, 0, 0)
    
    def body(text):
        pdf.set_font('Helvetica', '', 10)
        pdf.multi_cell(0, 5, text)
        pdf.cell(0, 2, '', ln=True)
    
    def bullet(text):
        pdf.set_font('Helvetica', '', 10)
        pdf.cell(5)
        pdf.cell(0, 5, f"- {text}", ln=True)
    
    # Cover Page
    pdf.add_page()
    pdf.set_font('Helvetica', '', 12)
    pdf.cell(0, 40, '', ln=True)
    pdf.set_font('Helvetica', 'B', 32)
    pdf.cell(0, 15, 'School CRM', align='C', ln=True)
    pdf.set_font('Helvetica', 'B', 20)
    pdf.cell(0, 12, 'User Manual', align='C', ln=True)
    pdf.set_font('Helvetica', '', 14)
    pdf.cell(0, 10, 'Complete Guide for All Users', align='C', ln=True)
    pdf.cell(0, 20, '', ln=True)
    pdf.set_font('Helvetica', '', 11)
    pdf.cell(0, 7, 'Multi-Tenant School Management & CRM System', align='C', ln=True)
    pdf.cell(0, 7, 'Version 1.0', align='C', ln=True)
    
    # Table of Contents
    pdf.add_page()
    section_title('Table of Contents')
    toc = [
        '1. Introduction & Overview',
        '2. System Requirements',
        '3. Getting Started - Login',
        '4. Super Admin Portal',
        '5. School Admin Guide',
        '6. Dashboard',
        '7. Student Management',
        '8. Staff Management',
        '9. Admissions & Leads (CRM)',
        '10. Fee Management',
        '11. Attendance Management',
        '12. Academics (Timetable, Exams, Results)',
        '13. Library Management',
        '14. Transport Management',
        '15. Hostel Management',
        '16. Canteen Management',
        '17. Sports Management',
        '18. Health Records',
        '19. Inventory Management',
        '20. Communication (SMS, Email, Notices)',
        '21. Reports & Analytics',
        '22. School Branding & Customization',
        '23. Data Import',
        '24. Parent Portal',
        '25. Role-Based Access Summary',
        '26. Troubleshooting & FAQ',
    ]
    for item in toc:
        body(item)
    
    # 1. Introduction
    pdf.add_page()
    section_title('1. Introduction & Overview')
    body('School CRM is a comprehensive, multi-tenant School Management and Customer Relationship Management (CRM) system. It is designed for schools, coaching institutes, and educational organizations to manage their entire operations digitally.')
    
    sub_title('Key Features')
    features = [
        'Multi-tenant architecture - One platform, multiple schools',
        'Role-based access control with 19 different user roles',
        'Complete student lifecycle management (Lead -> Admission -> Student)',
        'Fee collection and financial management',
        'Attendance tracking for students and staff',
        'Academic management (timetable, exams, results, report cards)',
        'Library, transport, hostel, canteen, sports management',
        'Health records and inventory tracking',
        'Communication via SMS, Email, WhatsApp, and notices',
        'Comprehensive reports and analytics',
        'School branding customization (logo, theme, login page)',
        'Subscription plan management for SaaS model',
        'Parent and student self-service portals',
        'Data import via Excel/CSV',
    ]
    for f in features:
        bullet(f)
    
    # 2. System Requirements
    pdf.add_page()
    section_title('2. System Requirements')
    sub_title('For Users (Browser)')
    bullet('Google Chrome (latest) - Recommended')
    bullet('Mozilla Firefox (latest)')
    bullet('Microsoft Edge (latest)')
    bullet('Stable internet connection')
    bullet('Works on Desktop, Tablet, and Mobile')
    
    pdf.cell(0, 5, '', ln=True)
    sub_title('For Server / Developer')
    bullet('Python 3.9+ with Flask framework')
    bullet('Node.js 16+ for React frontend')
    bullet('MySQL 8.0+ database')
    bullet('Minimum 2GB RAM, 20GB storage')
    
    # 3. Getting Started
    pdf.add_page()
    section_title('3. Getting Started - Login')
    body('The login page is the entry point for all users. Each user must enter their credentials to access the system.')
    
    sub_title('Login Steps for School Users')
    bullet('Step 1: Open the application URL in your browser')
    bullet('Step 2: Enter your School Code (provided by school admin)')
    bullet('Step 3: The login page will show your school branding (logo, name)')
    bullet('Step 4: Enter your Email address')
    bullet('Step 5: Enter your Password')
    bullet('Step 6: Click Login button')
    bullet('Step 7: You will be redirected to your dashboard based on your role')
    
    pdf.cell(0, 5, '', ln=True)
    sub_title('Login Steps for Super Admin')
    bullet('Step 1: Open the application URL')
    bullet('Step 2: Enter Email: admin@schoolcrm.com')
    bullet('Step 3: Enter Password: superadmin123')
    bullet('Step 4: School Code can be left empty or any value')
    bullet('Step 5: Click Login - redirected to Super Admin Portal')
    
    pdf.cell(0, 5, '', ln=True)
    sub_title('Default Login Credentials')
    pdf.set_font('Helvetica', 'B', 9)
    pdf.set_fill_color(240, 240, 240)
    pdf.cell(40, 7, 'Role', border=1, fill=True, align='C')
    pdf.cell(30, 7, 'School Code', border=1, fill=True, align='C')
    pdf.cell(70, 7, 'Email', border=1, fill=True, align='C')
    pdf.cell(30, 7, 'Password', border=1, fill=True, align='C')
    pdf.ln()
    pdf.set_font('Helvetica', '', 9)
    creds = [
        ('Super Admin', 'Any/Blank', 'admin@schoolcrm.com', 'superadmin123'),
        ('School Admin', 'Sikarwar', 'akhileshsingh241425@gmail.com', 'admin123'),
    ]
    for role, code, email, pw in creds:
        pdf.cell(40, 7, role, border=1, align='C')
        pdf.cell(30, 7, code, border=1, align='C')
        pdf.cell(70, 7, email, border=1, align='C')
        pdf.cell(30, 7, pw, border=1, align='C')
        pdf.ln()
    
    # 4. Super Admin Portal
    pdf.add_page()
    section_title('4. Super Admin Portal')
    body('The Super Admin Portal is the platform owner\'s control center. Only users with the super_admin role can access this section.')
    
    sub_title('4.1 Super Admin Dashboard')
    bullet('Total Schools: Number of schools registered on the platform')
    bullet('Active Schools: Schools with active subscriptions')
    bullet('Total Students: Sum of all students across all schools')
    bullet('Total Staff: Sum of all staff across all schools')
    bullet('Monthly Revenue: Total subscription revenue')
    bullet('Plan Distribution: Breakdown of schools by subscription plan')
    bullet('Recent Subscriptions: Latest subscription activities')
    
    pdf.cell(0, 5, '', ln=True)
    sub_title('4.2 Manage Schools')
    bullet('View list of all registered schools with search and filter')
    bullet('Add new school with admin credentials')
    bullet('Edit school details (name, code, contact)')
    bullet('Activate / Deactivate schools')
    bullet('Toggle individual features ON/OFF for each school')
    bullet('View school subscription status')
    
    pdf.cell(0, 5, '', ln=True)
    sub_title('4.3 Manage Plans')
    bullet('Create subscription plans (e.g., Basic, Standard, Premium)')
    bullet('Set monthly and yearly pricing')
    bullet('Define plan limits (max students, max staff)')
    bullet('Select which features are included in each plan')
    bullet('Activate / Deactivate plans')
    
    pdf.cell(0, 5, '', ln=True)
    sub_title('4.4 Manage Subscriptions')
    bullet('View all active and expired subscriptions')
    bullet('Assign subscription plans to schools')
    bullet('Track payment status (pending, paid, expired)')
    bullet('Set billing cycle (monthly / yearly)')
    
    # 5. School Admin Guide
    pdf.add_page()
    section_title('5. School Admin Guide')
    body('The School Admin has full control over their school. They can manage all modules, users, and settings within their school.')
    
    sub_title('5.1 First-Time Setup')
    bullet('Login with provided credentials')
    bullet('Go to Settings > School Branding to upload logo and customize theme')
    bullet('Set up fee structures in Fee Management')
    bullet('Import existing student data via Data Import')
    bullet('Create staff accounts and assign roles')
    bullet('Configure timetable and academic calendar')
    
    sub_title('5.2 User Management')
    bullet('Add new users (teachers, accountants, etc.) from Staff section')
    bullet('Assign appropriate roles to each user')
    bullet('Each role automatically gets specific module access')
    bullet('Deactivate users who leave the school')
    
    # 6-21: Module details
    modules_detail = [
        ('6. Dashboard', [
            'The dashboard provides an overview of key metrics at a glance.',
            'Shows total students, leads, fee collection, pending admissions.',
            'Quick links to frequently used modules.',
            'Recent activities and notifications.',
            'Different dashboard views based on user role.',
        ]),
        ('7. Student Management', [
            'Add new students with complete profile (name, class, section, parent info).',
            'View class-wise and section-wise student lists.',
            'Search students by name, admission number, or class.',
            'View individual student profile with all details.',
            'Manage student documents and certificates.',
            'Promote students to next class at year end.',
            'Transfer certificate generation.',
        ]),
        ('8. Staff Management', [
            'Add staff members with role assignment.',
            'Manage staff profiles (qualifications, experience, documents).',
            'Track staff attendance and leave.',
            'Department and designation management.',
            'Staff salary and payroll management.',
        ]),
        ('9. Admissions & Leads (CRM)', [
            'LEADS: Register new inquiries (walk-in, phone, website).',
            'Track lead source (newspaper, social media, referral, etc.).',
            'Follow-up management with notes and reminders.',
            'Lead status tracking (new, contacted, interested, converted, lost).',
            'ADMISSIONS: Convert leads to admission applications.',
            'Multi-stage admission pipeline (applied, document verification, approved, enrolled).',
            'Admission form with all required student details.',
            'Generate admission number on enrollment.',
        ]),
        ('10. Fee Management', [
            'Create fee structures (tuition, transport, hostel, exam, etc.).',
            'Class-wise and category-wise fee setup.',
            'Fee collection with receipt generation.',
            'Multiple payment modes (cash, online, cheque, UPI).',
            'Track pending fees and send reminders.',
            'Fee concession and discount management.',
            'Fee reports by class, date range, payment mode.',
        ]),
        ('11. Attendance Management', [
            'Mark daily attendance for students (class-wise).',
            'Mark staff attendance.',
            'Status options: Present, Absent, Late, Half-day, Leave.',
            'View attendance reports by student, class, or date range.',
            'Monthly attendance summary.',
            'SMS/notification to parents for absent students.',
        ]),
        ('12. Academics', [
            'TIMETABLE: Create class-wise period-wise timetable.',
            'CURRICULUM: Manage subjects, chapters, and syllabus.',
            'EXAMS: Create exam schedules with date, time, room allocation.',
            'RESULTS: Enter marks/grades subject-wise.',
            'REPORT CARDS: Auto-generate report cards with marks, grades, and remarks.',
            'Grade mapping and grading system configuration.',
            'Subject assignment to teachers.',
        ]),
        ('13. Library Management', [
            'Add and manage book records (title, author, ISBN, category).',
            'Issue books to students and staff.',
            'Return books with date tracking.',
            'Track overdue books and calculate fines.',
            'Book search and availability check.',
            'Library usage reports.',
        ]),
        ('14. Transport Management', [
            'Add bus routes with stops and timings.',
            'Manage vehicles (bus number, capacity, driver).',
            'Assign students to transport routes.',
            'Driver and conductor details management.',
            'Transport fee tracking.',
            'Route-wise student list generation.',
        ]),
        ('15. Hostel Management', [
            'Manage hostel buildings and rooms.',
            'Room allocation to students.',
            'Bed management and availability.',
            'Hostel fee tracking.',
            'Hostel attendance.',
            'Room change and vacancy management.',
        ]),
        ('16. Canteen Management', [
            'Manage daily menu items with prices.',
            'Process canteen orders.',
            'Track canteen sales and revenue.',
            'Canteen inventory management.',
            'Daily/monthly sales reports.',
        ]),
        ('17. Sports Management', [
            'Manage sports events and activities.',
            'Create teams and assign students.',
            'Record achievements and awards.',
            'Sports schedule management.',
            'Fitness record tracking.',
        ]),
        ('18. Health Records', [
            'Student health checkup records.',
            'Medical visit log (date, reason, treatment).',
            'Vaccination records.',
            'Allergy and medical condition tracking.',
            'Health reports and statistics.',
        ]),
        ('19. Inventory Management', [
            'Add and track school inventory items.',
            'Category-wise inventory organization.',
            'Stock in/out tracking.',
            'Low stock alerts.',
            'Procurement and purchase order management.',
            'Inventory usage reports.',
        ]),
        ('20. Communication', [
            'Send notices to all or selected groups.',
            'SMS notifications to parents.',
            'Email communication.',
            'WhatsApp integration for messaging.',
            'Class-wise and individual messaging.',
            'Communication history and logs.',
        ]),
        ('21. Reports & Analytics', [
            'ADMISSION REPORTS: Applications, enrolled, pending, by class.',
            'FEE REPORTS: Collection summary, pending, by payment mode.',
            'ATTENDANCE REPORTS: Daily, monthly, by class.',
            'MARKETING REPORTS: Lead conversion, source analysis.',
            'Custom date range filtering for all reports.',
            'Export reports for offline use.',
        ]),
    ]
    
    for title, details in modules_detail:
        if pdf.get_y() > 230:
            pdf.add_page()
        section_title(title)
        for d in details:
            bullet(d)
        pdf.cell(0, 3, '', ln=True)
    
    # 22. School Branding
    pdf.add_page()
    section_title('22. School Branding & Customization')
    body('School Admins can customize the look and feel of their school\'s CRM interface.')
    
    sub_title('Available Customizations')
    bullet('School Logo: Upload your school logo (displayed in sidebar and login page)')
    bullet('Login Background Image: Custom background image for the login page')
    bullet('Banner Image: Banner for school dashboard/header')
    bullet('School Tagline: A tagline displayed on the login page')
    bullet('Theme Color: Primary color for the UI theme')
    
    pdf.cell(0, 5, '', ln=True)
    sub_title('How to Update Branding')
    bullet('Step 1: Login as School Admin')
    bullet('Step 2: Go to sidebar > School Branding')
    bullet('Step 3: Upload images using the upload buttons')
    bullet('Step 4: Enter tagline and select theme color')
    bullet('Step 5: Click Save to apply changes')
    bullet('Step 6: Changes reflect on the login page immediately')
    
    # 23. Data Import
    if pdf.get_y() > 200:
        pdf.add_page()
    section_title('23. Data Import')
    body('Bulk import data using Excel or CSV files to quickly set up student and staff records.')
    bullet('Navigate to Data Import from the sidebar')
    bullet('Download the sample template for the data type')
    bullet('Fill in the data following the template format')
    bullet('Upload the filled file')
    bullet('System validates data and shows errors if any')
    bullet('Confirm to import valid records')
    
    # 24. Parent Portal
    pdf.add_page()
    section_title('24. Parent Portal')
    body('Parents have a dedicated portal to track their child\'s progress and communicate with school.')
    
    sub_title('Parent Features')
    bullet('View child\'s attendance (daily and monthly)')
    bullet('Check exam results and report cards')
    bullet('View fee payment status and pending dues')
    bullet('Receive school notices and announcements')
    bullet('Communicate with class teacher')
    bullet('View timetable and academic calendar')
    bullet('Access multiple children from one account')
    
    sub_title('Parent Login')
    bullet('Parents receive login credentials from school')
    bullet('Login with School Code + Email/Phone + Password')
    bullet('Redirected to "My Children" page after login')
    
    # 25. Role Summary
    pdf.add_page()
    section_title('25. Role-Based Access Summary')
    body('The system has 19 different roles. Each role has specific module access.')
    
    pdf.set_font('Helvetica', 'B', 8)
    pdf.set_fill_color(44, 62, 80)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(35, 7, 'Role', border=1, fill=True, align='C')
    pdf.cell(150, 7, 'Accessible Modules', border=1, fill=True, align='C')
    pdf.ln()
    pdf.set_text_color(0, 0, 0)
    
    role_summary = {
        'Super Admin': 'ALL modules + Super Admin Portal (Schools, Plans, Subscriptions)',
        'School Admin': 'ALL school modules + Settings + Branding + Data Import',
        'Principal': 'ALL school modules (similar to School Admin)',
        'Teacher': 'Dashboard, Students, Attendance, Academics, Library, Communication, Parents, Reports',
        'Accountant': 'Dashboard, Students, Fees, Inventory, Reports',
        'Counselor': 'Dashboard, Students, Admissions, Leads, Parents, Communication',
        'Receptionist': 'Dashboard, Students, Admissions, Leads, Parents, Communication',
        'Exam Controller': 'Dashboard, Students, Academics, Reports',
        'HOD': 'Dashboard, Students, Attendance, Academics, Library, Communication, Reports',
        'HR Manager': 'Dashboard, Staff, Attendance, Communication, Reports',
        'Hostel Warden': 'Dashboard, Students, Hostel, Attendance, Health, Communication',
        'Transport Mgr': 'Dashboard, Students, Transport, Communication, Reports',
        'Librarian': 'Dashboard, Students, Library, Reports',
        'Canteen Mgr': 'Dashboard, Canteen, Inventory, Reports',
        'Sports PTI': 'Dashboard, Students, Sports, Health, Communication',
        'Health Officer': 'Dashboard, Students, Health, Communication, Reports',
        'Lab Assistant': 'Dashboard, Academics, Inventory',
        'Parent': 'Dashboard (My Children), Communication',
        'Student': 'Dashboard (limited self-service)',
    }
    
    pdf.set_font('Helvetica', '', 7)
    for i, (role, modules) in enumerate(role_summary.items()):
        if i % 2 == 0:
            pdf.set_fill_color(245, 245, 245)
        else:
            pdf.set_fill_color(255, 255, 255)
        pdf.cell(35, 6, role, border=1, fill=True)
        pdf.cell(150, 6, modules, border=1, fill=True)
        pdf.ln()
    
    # 26. Troubleshooting
    pdf.add_page()
    section_title('26. Troubleshooting & FAQ')
    
    faqs = [
        ('Q: I forgot my password. What should I do?',
         'A: Contact your School Admin. They can reset your password from the Staff/User management section.'),
        ('Q: I cannot see a module in my sidebar.',
         'A: Your role may not have access to that module. Contact your School Admin to check your role and permissions.'),
        ('Q: Fee receipt is not generating.',
         'A: Ensure all required fields are filled. Check if the fee structure is properly configured for the class.'),
        ('Q: Attendance is not saving.',
         'A: Make sure you have selected the correct class and date. Check your internet connection and try again.'),
        ('Q: Login page shows "Invalid school code".',
         'A: Verify the school code with your administrator. The code is case-sensitive.'),
        ('Q: Super Admin cannot see schools.',
         'A: Ensure you are logged in with the super_admin role. Check if schools exist in the database.'),
        ('Q: Images are not uploading in School Branding.',
         'A: Ensure the file is in PNG, JPG, JPEG, GIF, WEBP, or SVG format. Maximum recommended size is 5MB.'),
        ('Q: How to add a new role/user?',
         'A: Go to Staff section > Add Staff > Select the appropriate role during creation.'),
        ('Q: Can I use this on mobile?',
         'A: Yes, the interface is responsive and works on mobile browsers. No separate app is needed.'),
    ]
    
    for q, a in faqs:
        if pdf.get_y() > 260:
            pdf.add_page()
        sub_title(q)
        body(a)
    
    filepath = os.path.join(OUTPUT_DIR, 'School_CRM_User_Manual.pdf')
    pdf.output(filepath)
    print(f'[OK] User Manual PDF: {filepath}')


# ======================== MAIN ========================

if __name__ == '__main__':
    print('Generating documentation...')
    print(f'Output directory: {OUTPUT_DIR}')
    print()
    create_flowchart_pdf()
    create_roles_excel()
    create_user_manual()
    print()
    print('All documents generated successfully!')
    print(f'Check: {OUTPUT_DIR}')
