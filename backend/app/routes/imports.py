"""
Bulk Import System - Data Migration Support
Handles CSV upload, validation, preview, and import for all modules.
Supports school migration from other software.
"""

from flask import Blueprint, request, g, Response, send_file
from app import db
from app.models.student import Student, Class, Section, AcademicYear, ParentDetail
from app.models.staff import Staff
from app.models.academic import Subject, Syllabus, AcademicCalendar
from app.models.fee import FeeCategory, FeeStructure
from app.models.library import LibraryBook, LibraryCategory
from app.utils.decorators import school_required, role_required
from app.utils.helpers import success_response, error_response
from datetime import datetime, date
import csv
import io
import json
import hashlib
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import msoffcrypto

imports_bp = Blueprint('imports', __name__)


# ============================================================
# CSV TEMPLATES - Download sample templates
# ============================================================

TEMPLATES = {
    'students': {
        'headers': ['admission_no', 'roll_no', 'first_name', 'last_name', 'gender', 'date_of_birth',
                     'class_name', 'section_name', 'blood_group', 'religion', 'category', 'nationality',
                     'aadhar_no', 'address', 'city', 'state', 'pincode',
                     'mother_tongue', 'emergency_contact', 'emergency_person',
                     'medical_conditions', 'allergies', 'previous_school', 'transport_mode',
                     'father_name', 'father_phone', 'father_email', 'father_occupation',
                     'mother_name', 'mother_phone', 'mother_email', 'mother_occupation'],
        'sample': ['ADM001', '1', 'Rahul', 'Sharma', 'male', '2015-05-10',
                    'Class 5', 'A', 'B+', 'Hindu', 'General', 'Indian',
                    '123456789012', '123 Main St', 'Delhi', 'Delhi', '110001',
                    'Hindi', '9876543210', 'Father', '', '', 'Previous School', 'bus',
                    'Rajesh Sharma', '9876543210', 'rajesh@email.com', 'Business',
                    'Priya Sharma', '9876543211', 'priya@email.com', 'Teacher'],
        'description': 'Student data with parent info. class_name and section_name must match existing classes.'
    },
    'staff': {
        'headers': ['employee_id', 'first_name', 'last_name', 'gender', 'date_of_birth',
                     'phone', 'email', 'qualification', 'experience_years', 'designation',
                     'department', 'date_of_joining', 'salary', 'address', 'city', 'state',
                     'aadhar_no', 'pan_no', 'bank_name', 'bank_account_no', 'ifsc_code',
                     'staff_type', 'contract_type', 'blood_group', 'emergency_contact', 'emergency_person',
                     'marital_status'],
        'sample': ['EMP001', 'Anita', 'Verma', 'female', '1985-03-15',
                    '9876543210', 'anita@school.com', 'M.Ed', '10', 'Senior Teacher',
                    'Science', '2020-04-01', '45000', '456 Oak St', 'Mumbai', 'Maharashtra',
                    '987654321012', 'ABCDE1234F', 'SBI', '12345678901234', 'SBIN0001234',
                    'teaching', 'permanent', 'O+', '9876543211', 'Husband',
                    'married'],
        'description': 'Staff/Teachers data. staff_type: teaching/non_teaching/admin. contract_type: permanent/contract/probation/part_time'
    },
    'subjects': {
        'headers': ['name', 'code', 'type', 'description', 'credit_hours', 'is_elective'],
        'sample': ['Mathematics', 'MATH', 'theory', 'Core mathematics', '5', 'false'],
        'description': 'Subject master list. type: theory/practical/both. is_elective: true/false'
    },
    'class_subjects': {
        'headers': ['class_name', 'subject_code', 'teacher_employee_id'],
        'sample': ['Class 5', 'MATH', 'EMP001'],
        'description': 'Which subject in which class with teacher. Uses class name and subject code to match.'
    },
    'syllabus': {
        'headers': ['class_name', 'subject_code', 'chapter_number', 'chapter_name', 'topics',
                     'learning_objectives', 'estimated_hours', 'term'],
        'sample': ['Class 5', 'MATH', '1', 'Numbers and Numeration', 'Place value, Indian system, International system',
                    'Understand large numbers, Compare numbers', '8', 'term1'],
        'description': 'Chapter-wise syllabus. term: term1/term2/term3/annual'
    },
    'fee_structure': {
        'headers': ['class_name', 'category_name', 'amount', 'frequency', 'due_date',
                     'late_fee_amount', 'late_fee_type', 'grace_period_days'],
        'sample': ['Class 5', 'Tuition Fee', '5000', 'monthly', '2026-04-10',
                    '50', 'fixed', '7'],
        'description': 'Fee structure per class. frequency: one_time/monthly/quarterly/half_yearly/yearly. late_fee_type: fixed/percentage/per_day. Creates category if not exists.'
    },
    'library_books': {
        'headers': ['title', 'author', 'isbn', 'publisher', 'edition', 'language', 'subject',
                     'publication_year', 'pages', 'total_copies', 'rack_no', 'price', 'category_name'],
        'sample': ['NCERT Mathematics 5', 'NCERT', '978-81-7450-123-4', 'NCERT', '2024', 'English', 'Mathematics',
                    '2024', '200', '10', 'A1-S3', '150', 'Textbooks'],
        'description': 'Library books. Creates category if not exists.'
    },
    'calendar_events': {
        'headers': ['title', 'description', 'event_type', 'start_date', 'end_date',
                     'is_holiday', 'applies_to', 'class_name'],
        'sample': ['Diwali Vacation', 'School closed for Diwali', 'vacation', '2026-10-20', '2026-10-30',
                    'true', 'all', ''],
        'description': 'Academic calendar events. event_type: holiday/exam/ptm/event/cultural/sports/meeting/deadline/vacation/other. applies_to: all/students/staff/specific_class'
    },
    'timetable': {
        'headers': ['class_name', 'section_name', 'subject_code', 'teacher_employee_id',
                     'day_of_week', 'period_number', 'start_time', 'end_time', 'room_no'],
        'sample': ['Class 5', 'A', 'MATH', 'EMP001', 'monday', '1', '08:00', '08:45', 'Room 5A'],
        'description': 'Timetable entries. day_of_week: monday/tuesday/wednesday/thursday/friday/saturday'
    },
}


@imports_bp.route('/templates', methods=['GET'])
@school_required
def list_templates():
    """List all available import templates"""
    result = []
    for key, tmpl in TEMPLATES.items():
        result.append({
            'key': key,
            'name': key.replace('_', ' ').title(),
            'description': tmpl['description'],
            'columns': len(tmpl['headers']),
            'headers': tmpl['headers'],
        })
    return success_response(result)


@imports_bp.route('/templates/<template_key>/download', methods=['GET'])
@school_required
def download_template(template_key):
    """Download Excel template with formatted headers, sample row, and password protection"""
    tmpl = TEMPLATES.get(template_key)
    if not tmpl:
        return error_response('Template not found', 404)

    # Generate password from school_id
    password = generate_template_password(g.school_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = template_key.replace('_', ' ').title()

    # Color scheme per module
    MODULE_HEADER_COLORS = {
        'students': '1565C0',
        'staff': '2E7D32',
        'subjects': 'E65100',
        'class_subjects': '7B1FA2',
        'syllabus': '0277BD',
        'fee_structure': 'C62828',
        'library_books': '6A1B9A',
        'calendar_events': 'AD1457',
        'timetable': '00695C',
    }
    header_color = MODULE_HEADER_COLORS.get(template_key, '1565C0')

    # Styles
    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='BDBDBD'),
        right=Side(style='thin', color='BDBDBD'),
        top=Side(style='thin', color='BDBDBD'),
        bottom=Side(style='thin', color='BDBDBD'),
    )
    sample_fill = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')
    sample_font = Font(name='Calibri', color='795548', italic=True, size=10)
    data_font = Font(name='Calibri', size=10)

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(tmpl['headers']))
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = f"📋 {template_key.replace('_', ' ').title()} Import Template"
    title_cell.font = Font(name='Calibri', bold=True, size=14, color=header_color)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35

    # Description row
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(tmpl['headers']))
    desc_cell = ws.cell(row=2, column=1)
    desc_cell.value = tmpl['description']
    desc_cell.font = Font(name='Calibri', italic=True, size=9, color='757575')
    desc_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[2].height = 30

    # Header row (row 3)
    for col_idx, header in enumerate(tmpl['headers'], 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    ws.row_dimensions[3].height = 28

    # Sample data row (row 4) - styled differently to show it's example data
    for col_idx, value in enumerate(tmpl['sample'], 1):
        cell = ws.cell(row=4, column=col_idx, value=value)
        cell.font = sample_font
        cell.fill = sample_fill
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = thin_border

    # Empty rows for data entry (rows 5-24)
    for row in range(5, 25):
        for col_idx in range(1, len(tmpl['headers']) + 1):
            cell = ws.cell(row=row, column=col_idx, value='')
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')

    # Auto-width columns (based on header + sample data)
    for col_idx, header in enumerate(tmpl['headers'], 1):
        sample_val = tmpl['sample'][col_idx - 1] if col_idx <= len(tmpl['sample']) else ''
        max_len = max(len(str(header)), len(str(sample_val)), 10)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 35)

    # Freeze header row
    ws.freeze_panes = 'A4'

    # Auto-filter on header row
    ws.auto_filter.ref = f"A3:{get_column_letter(len(tmpl['headers']))}3"

    # Instructions sheet
    ws_info = wb.create_sheet('Instructions')
    ws_info.sheet_properties.tabColor = 'FF9800'

    instructions = [
        ['📌 Import Instructions'],
        [''],
        ['1. Fill your data starting from row 5 (row 4 is a sample - you can delete or overwrite it)'],
        ['2. Do NOT change the header row (row 3) - column names must match exactly'],
        ['3. Date format: YYYY-MM-DD (e.g., 2025-05-15)'],
        ['4. Save the file as .xlsx (Excel format)'],
        ['5. Upload in the Data Import section of School CRM'],
        [''],
        [f'Template: {template_key.replace("_", " ").title()}'],
        [f'Total Columns: {len(tmpl["headers"])}'],
        [f'Description: {tmpl["description"]}'],
        [''],
        ['Column Details:'],
    ]
    for h in tmpl['headers']:
        instructions.append([f'  • {h}'])

    for row_idx, row_data in enumerate(instructions, 1):
        cell = ws_info.cell(row=row_idx, column=1, value=row_data[0])
        if row_idx == 1:
            cell.font = Font(bold=True, size=14, color=header_color)
        elif row_data[0].startswith('  •'):
            cell.font = Font(size=10, color='424242')
        else:
            cell.font = Font(size=10)

    ws_info.column_dimensions['A'].width = 80

    # Save to bytes
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    # Encrypt with password using msoffcrypto
    encrypted_buffer = io.BytesIO()
    file_obj = msoffcrypto.OfficeFile(excel_buffer)
    file_obj.load_key(password=password)
    file_obj.encrypt(password=password, outfile=encrypted_buffer)
    encrypted_buffer.seek(0)

    return send_file(
        encrypted_buffer,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'{template_key}_template.xlsx'
    )


def generate_template_password(school_id):
    """Generate a consistent password for template files based on school ID"""
    raw = f"SchoolCRM_{school_id}_templates_2025"
    return hashlib.sha256(raw.encode()).hexdigest()[:12]


@imports_bp.route('/template-password', methods=['GET'])
@school_required
def get_template_password():
    """Get the password for template Excel files"""
    password = generate_template_password(g.school_id)
    return success_response({'password': password})


# ============================================================
# VALIDATE & PREVIEW - Parse CSV and show preview with errors
# ============================================================

def parse_csv_data(file_content):
    """Parse CSV content and return rows as list of dicts"""
    try:
        content = file_content.decode('utf-8-sig')  # Handle BOM
    except UnicodeDecodeError:
        content = file_content.decode('latin-1')

    reader = csv.DictReader(io.StringIO(content))
    rows = []
    for i, row in enumerate(reader):
        # Strip whitespace from keys and values
        cleaned = {k.strip(): v.strip() if v else '' for k, v in row.items() if k}
        cleaned['_row_num'] = i + 2  # 1-indexed, +1 for header
        rows.append(cleaned)
    return rows, list(reader.fieldnames) if reader.fieldnames else []


def parse_excel_data(file_content):
    """Parse Excel (.xlsx) content and return rows as list of dicts"""
    wb = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
    ws = wb.active

    # Find header row (row 3 in our template, but check for headers)
    headers = []
    header_row = 1
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), 1):
        # Check if this row looks like headers (mostly non-empty strings)
        non_empty = [c for c in row if c and str(c).strip()]
        if len(non_empty) >= 3 and all(isinstance(c, str) for c in non_empty):
            headers = [str(c).strip() if c else '' for c in row]
            header_row = row_idx
            break

    if not headers:
        return [], []

    rows = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), header_row + 1):
        values = [str(c).strip() if c is not None else '' for c in row]
        # Skip completely empty rows
        if not any(v for v in values):
            continue
        cleaned = {}
        for i, header in enumerate(headers):
            if header and i < len(values):
                cleaned[header] = values[i]
        cleaned['_row_num'] = row_idx
        rows.append(cleaned)

    wb.close()
    return rows, [h for h in headers if h]


def validate_students(rows, school_id):
    """Validate student import data"""
    errors = []
    valid_rows = []
    # Pre-fetch classes and sections for lookup
    classes = {c.name.lower(): c for c in Class.query.filter_by(school_id=school_id).all()}
    sections = {}
    for s in Section.query.filter(Section.class_id.in_([c.id for c in classes.values()])).all():
        key = f"{s.class_id}_{s.name.lower()}"
        sections[key] = s

    existing_admissions = set(
        s.admission_no.lower() for s in Student.query.filter_by(school_id=school_id).all() if s.admission_no
    )

    for row in rows:
        row_errors = []
        rn = row['_row_num']

        if not row.get('first_name'):
            row_errors.append('first_name is required')
        if not row.get('class_name'):
            row_errors.append('class_name is required')
        if row.get('admission_no') and row['admission_no'].lower() in existing_admissions:
            row_errors.append(f"admission_no '{row['admission_no']}' already exists")
        if row.get('class_name') and row['class_name'].lower() not in classes:
            row_errors.append(f"class '{row['class_name']}' not found. Create it first.")
        if row.get('gender') and row['gender'].lower() not in ('male', 'female', 'other'):
            row_errors.append(f"invalid gender '{row['gender']}'. Use: male/female/other")
        if row.get('date_of_birth'):
            try:
                datetime.strptime(row['date_of_birth'], '%Y-%m-%d')
            except ValueError:
                row_errors.append('date_of_birth must be YYYY-MM-DD format')

        # Validate section if class exists
        if row.get('section_name') and row.get('class_name') and row['class_name'].lower() in classes:
            cls = classes[row['class_name'].lower()]
            sec_key = f"{cls.id}_{row['section_name'].lower()}"
            if sec_key not in sections:
                row_errors.append(f"section '{row['section_name']}' not found in class '{row['class_name']}'")

        if row_errors:
            errors.append({'row': rn, 'errors': row_errors, 'data': {k: v for k, v in row.items() if k != '_row_num'}})
        else:
            valid_rows.append(row)

    return valid_rows, errors


def validate_staff(rows, school_id):
    """Validate staff import data"""
    errors = []
    valid_rows = []
    existing_emp_ids = set(
        s.employee_id.lower() for s in Staff.query.filter_by(school_id=school_id).all() if s.employee_id
    )

    for row in rows:
        row_errors = []
        rn = row['_row_num']

        if not row.get('first_name'):
            row_errors.append('first_name is required')
        if not row.get('employee_id'):
            row_errors.append('employee_id is required')
        elif row['employee_id'].lower() in existing_emp_ids:
            row_errors.append(f"employee_id '{row['employee_id']}' already exists")
        if row.get('gender') and row['gender'].lower() not in ('male', 'female', 'other'):
            row_errors.append('invalid gender')
        if row.get('staff_type') and row['staff_type'].lower() not in ('teaching', 'non_teaching', 'admin'):
            row_errors.append('staff_type must be: teaching/non_teaching/admin')
        if row.get('date_of_joining'):
            try:
                datetime.strptime(row['date_of_joining'], '%Y-%m-%d')
            except ValueError:
                row_errors.append('date_of_joining must be YYYY-MM-DD')

        if row_errors:
            errors.append({'row': rn, 'errors': row_errors, 'data': {k: v for k, v in row.items() if k != '_row_num'}})
        else:
            valid_rows.append(row)

    return valid_rows, errors


def validate_subjects(rows, school_id):
    """Validate subjects import"""
    errors = []
    valid_rows = []
    existing = set(s.name.lower() for s in Subject.query.filter_by(school_id=school_id).all())

    for row in rows:
        row_errors = []
        if not row.get('name'):
            row_errors.append('name is required')
        elif row['name'].lower() in existing:
            row_errors.append(f"subject '{row['name']}' already exists")
        if row.get('type') and row['type'].lower() not in ('theory', 'practical', 'both'):
            row_errors.append('type must be: theory/practical/both')

        if row_errors:
            errors.append({'row': row['_row_num'], 'errors': row_errors, 'data': {k: v for k, v in row.items() if k != '_row_num'}})
        else:
            valid_rows.append(row)

    return valid_rows, errors


def validate_syllabus(rows, school_id):
    """Validate syllabus import"""
    errors = []
    valid_rows = []
    classes = {c.name.lower(): c for c in Class.query.filter_by(school_id=school_id).all()}
    subjects = {s.code.lower() if s.code else s.name.lower(): s for s in Subject.query.filter_by(school_id=school_id).all()}

    for row in rows:
        row_errors = []
        if not row.get('class_name'):
            row_errors.append('class_name is required')
        elif row['class_name'].lower() not in classes:
            row_errors.append(f"class '{row['class_name']}' not found")
        if not row.get('subject_code'):
            row_errors.append('subject_code is required')
        elif row['subject_code'].lower() not in subjects:
            row_errors.append(f"subject '{row['subject_code']}' not found")
        if not row.get('chapter_name'):
            row_errors.append('chapter_name is required')
        if row.get('term') and row['term'].lower() not in ('term1', 'term2', 'term3', 'annual'):
            row_errors.append('term must be: term1/term2/term3/annual')

        if row_errors:
            errors.append({'row': row['_row_num'], 'errors': row_errors, 'data': {k: v for k, v in row.items() if k != '_row_num'}})
        else:
            valid_rows.append(row)

    return valid_rows, errors


def validate_fee_structure(rows, school_id):
    """Validate fee structure import"""
    errors = []
    valid_rows = []
    classes = {c.name.lower(): c for c in Class.query.filter_by(school_id=school_id).all()}

    for row in rows:
        row_errors = []
        if not row.get('class_name'):
            row_errors.append('class_name is required')
        elif row['class_name'].lower() not in classes:
            row_errors.append(f"class '{row['class_name']}' not found")
        if not row.get('category_name'):
            row_errors.append('category_name is required')
        if not row.get('amount'):
            row_errors.append('amount is required')
        elif not row['amount'].replace('.', '').isdigit():
            row_errors.append('amount must be a number')
        if row.get('frequency') and row['frequency'].lower() not in ('one_time', 'monthly', 'quarterly', 'half_yearly', 'yearly'):
            row_errors.append('frequency must be: one_time/monthly/quarterly/half_yearly/yearly')

        if row_errors:
            errors.append({'row': row['_row_num'], 'errors': row_errors, 'data': {k: v for k, v in row.items() if k != '_row_num'}})
        else:
            valid_rows.append(row)

    return valid_rows, errors


def validate_library_books(rows, school_id):
    """Validate library books import"""
    errors = []
    valid_rows = []

    for row in rows:
        row_errors = []
        if not row.get('title'):
            row_errors.append('title is required')
        if not row.get('author'):
            row_errors.append('author is required')
        if row.get('total_copies') and not row['total_copies'].isdigit():
            row_errors.append('total_copies must be a number')

        if row_errors:
            errors.append({'row': row['_row_num'], 'errors': row_errors, 'data': {k: v for k, v in row.items() if k != '_row_num'}})
        else:
            valid_rows.append(row)

    return valid_rows, errors


def validate_calendar_events(rows, school_id):
    """Validate calendar events import"""
    errors = []
    valid_rows = []
    valid_types = ('holiday', 'exam', 'ptm', 'event', 'cultural', 'sports', 'meeting', 'deadline', 'vacation', 'other')

    for row in rows:
        row_errors = []
        if not row.get('title'):
            row_errors.append('title is required')
        if not row.get('start_date'):
            row_errors.append('start_date is required')
        else:
            try:
                datetime.strptime(row['start_date'], '%Y-%m-%d')
            except ValueError:
                row_errors.append('start_date must be YYYY-MM-DD')
        if row.get('event_type') and row['event_type'].lower() not in valid_types:
            row_errors.append(f"event_type must be one of: {', '.join(valid_types)}")

        if row_errors:
            errors.append({'row': row['_row_num'], 'errors': row_errors, 'data': {k: v for k, v in row.items() if k != '_row_num'}})
        else:
            valid_rows.append(row)

    return valid_rows, errors


VALIDATORS = {
    'students': validate_students,
    'staff': validate_staff,
    'subjects': validate_subjects,
    'syllabus': validate_syllabus,
    'fee_structure': validate_fee_structure,
    'library_books': validate_library_books,
    'calendar_events': validate_calendar_events,
}


@imports_bp.route('/validate/<import_type>', methods=['POST'])
@role_required('school_admin')
def validate_import(import_type):
    """Upload CSV, validate, and return preview with errors"""
    if import_type not in VALIDATORS:
        return error_response(f'Invalid import type: {import_type}', 400)

    if 'file' not in request.files:
        return error_response('No file uploaded', 400)

    file = request.files['file']
    if not file.filename.endswith('.csv') and not file.filename.endswith('.xlsx'):
        return error_response('Only CSV or Excel (.xlsx) files are supported.', 400)

    content = file.read()
    if len(content) > 10 * 1024 * 1024:  # 10MB limit
        return error_response('File too large. Maximum 10MB.', 400)

    if file.filename.endswith('.xlsx'):
        rows, headers = parse_excel_data(content)
    else:
        rows, headers = parse_csv_data(content)
    if not rows:
        return error_response('No data rows found in CSV', 400)

    # Validate expected headers
    expected = set(TEMPLATES[import_type]['headers'])
    found = set(h.strip() for h in headers if h)
    missing = expected - found
    extra = found - expected - {'_row_num'}

    valid_rows, errors = VALIDATORS[import_type](rows, g.school_id)

    return success_response({
        'total_rows': len(rows),
        'valid_rows': len(valid_rows),
        'error_rows': len(errors),
        'errors': errors[:50],  # Limit error display
        'preview': [
            {k: v for k, v in r.items() if k != '_row_num'}
            for r in valid_rows[:20]
        ],  # First 20 valid rows as preview
        'missing_headers': list(missing),
        'extra_headers': list(extra),
        'headers_found': list(found),
    })


# ============================================================
# IMPORT - Actually import valid data
# ============================================================

@imports_bp.route('/execute/<import_type>', methods=['POST'])
@role_required('school_admin')
def execute_import(import_type):
    """Execute the actual import after validation"""
    if import_type not in VALIDATORS:
        return error_response(f'Invalid import type: {import_type}', 400)

    if 'file' not in request.files:
        return error_response('No file uploaded', 400)

    file = request.files['file']
    content = file.read()
    if file.filename.endswith('.xlsx'):
        rows, headers = parse_excel_data(content)
    else:
        rows, headers = parse_csv_data(content)

    if not rows:
        return error_response('No data rows found', 400)

    valid_rows, errors = VALIDATORS[import_type](rows, g.school_id)

    if not valid_rows:
        return error_response('No valid rows to import', 400)

    try:
        if import_type == 'students':
            imported = import_students(valid_rows, g.school_id)
        elif import_type == 'staff':
            imported = import_staff(valid_rows, g.school_id)
        elif import_type == 'subjects':
            imported = import_subjects(valid_rows, g.school_id)
        elif import_type == 'syllabus':
            imported = import_syllabus(valid_rows, g.school_id)
        elif import_type == 'fee_structure':
            imported = import_fee_structure(valid_rows, g.school_id)
        elif import_type == 'library_books':
            imported = import_library_books(valid_rows, g.school_id)
        elif import_type == 'calendar_events':
            imported = import_calendar_events(valid_rows, g.school_id)
        else:
            return error_response('Not implemented', 400)

        db.session.commit()
        return success_response({
            'imported': imported,
            'skipped': len(errors),
            'total': len(rows),
        }, f'Successfully imported {imported} records')

    except Exception as e:
        db.session.rollback()
        return error_response(f'Import failed: {str(e)}', 500)


# ============================================================
# IMPORT FUNCTIONS
# ============================================================

def import_students(rows, school_id):
    """Import students with parent info"""
    classes = {c.name.lower(): c for c in Class.query.filter_by(school_id=school_id).all()}
    sections = {}
    for s in Section.query.filter(Section.class_id.in_([c.id for c in classes.values()])).all():
        sections[f"{s.class_id}_{s.name.lower()}"] = s

    academic_year = AcademicYear.query.filter_by(school_id=school_id, is_current=True).first()
    count = 0

    for row in rows:
        cls = classes.get(row.get('class_name', '').lower())
        if not cls:
            continue

        sec = None
        if row.get('section_name'):
            sec = sections.get(f"{cls.id}_{row['section_name'].lower()}")

        dob = None
        if row.get('date_of_birth'):
            try:
                dob = datetime.strptime(row['date_of_birth'], '%Y-%m-%d').date()
            except ValueError:
                pass

        student = Student(
            school_id=school_id,
            admission_no=row.get('admission_no') or None,
            roll_no=row.get('roll_no') or None,
            first_name=row['first_name'],
            last_name=row.get('last_name', ''),
            gender=row.get('gender', 'male').lower(),
            date_of_birth=dob,
            blood_group=row.get('blood_group') or None,
            religion=row.get('religion') or None,
            category=row.get('category') or None,
            nationality=row.get('nationality', 'Indian'),
            aadhar_no=row.get('aadhar_no') or None,
            address=row.get('address') or None,
            city=row.get('city') or None,
            state=row.get('state') or None,
            pincode=row.get('pincode') or None,
            current_class_id=cls.id,
            current_section_id=sec.id if sec else None,
            academic_year_id=academic_year.id if academic_year else None,
            mother_tongue=row.get('mother_tongue') or None,
            emergency_contact=row.get('emergency_contact') or None,
            emergency_person=row.get('emergency_person') or None,
            medical_conditions=row.get('medical_conditions') or None,
            allergies=row.get('allergies') or None,
            previous_school=row.get('previous_school') or None,
            transport_mode=row.get('transport_mode') or None,
            status='active',
            admission_date=date.today(),
        )
        db.session.add(student)
        db.session.flush()  # Get student ID

        # Create parent details if father/mother info provided
        if row.get('father_name'):
            father = ParentDetail(
                school_id=school_id,
                student_id=student.id,
                relation='father',
                name=row['father_name'],
                phone=row.get('father_phone') or None,
                email=row.get('father_email') or None,
                occupation=row.get('father_occupation') or None,
            )
            db.session.add(father)

        if row.get('mother_name'):
            mother = ParentDetail(
                school_id=school_id,
                student_id=student.id,
                relation='mother',
                name=row['mother_name'],
                phone=row.get('mother_phone') or None,
                email=row.get('mother_email') or None,
                occupation=row.get('mother_occupation') or None,
            )
            db.session.add(mother)

        count += 1

    return count


def import_staff(rows, school_id):
    """Import staff members"""
    count = 0
    for row in rows:
        dob = None
        if row.get('date_of_birth'):
            try:
                dob = datetime.strptime(row['date_of_birth'], '%Y-%m-%d').date()
            except ValueError:
                pass

        doj = None
        if row.get('date_of_joining'):
            try:
                doj = datetime.strptime(row['date_of_joining'], '%Y-%m-%d').date()
            except ValueError:
                pass

        staff = Staff(
            school_id=school_id,
            employee_id=row['employee_id'],
            first_name=row['first_name'],
            last_name=row.get('last_name', ''),
            gender=row.get('gender', 'male').lower(),
            date_of_birth=dob,
            phone=row.get('phone') or None,
            email=row.get('email') or None,
            qualification=row.get('qualification') or None,
            experience_years=int(row['experience_years']) if row.get('experience_years', '').isdigit() else None,
            designation=row.get('designation') or None,
            department=row.get('department') or None,
            date_of_joining=doj or date.today(),
            salary=float(row['salary']) if row.get('salary', '').replace('.', '').isdigit() else None,
            address=row.get('address') or None,
            city=row.get('city') or None,
            state=row.get('state') or None,
            aadhar_no=row.get('aadhar_no') or None,
            pan_no=row.get('pan_no') or None,
            bank_name=row.get('bank_name') or None,
            bank_account_no=row.get('bank_account_no') or None,
            ifsc_code=row.get('ifsc_code') or None,
            staff_type=row.get('staff_type', 'teaching').lower(),
            contract_type=row.get('contract_type', 'permanent').lower(),
            blood_group=row.get('blood_group') or None,
            emergency_contact=row.get('emergency_contact') or None,
            emergency_person=row.get('emergency_person') or None,
            marital_status=row.get('marital_status') or None,
            status='active',
        )
        db.session.add(staff)
        count += 1

    return count


def import_subjects(rows, school_id):
    """Import subjects"""
    count = 0
    for row in rows:
        subject = Subject(
            school_id=school_id,
            name=row['name'],
            code=row.get('code') or None,
            type=row.get('type', 'theory').lower(),
            description=row.get('description') or None,
            credit_hours=float(row['credit_hours']) if row.get('credit_hours', '').replace('.', '').isdigit() else None,
            is_elective=row.get('is_elective', 'false').lower() == 'true',
            is_active=True,
        )
        db.session.add(subject)
        count += 1

    return count


def import_syllabus(rows, school_id):
    """Import syllabus chapters"""
    classes = {c.name.lower(): c for c in Class.query.filter_by(school_id=school_id).all()}
    subjects = {}
    for s in Subject.query.filter_by(school_id=school_id).all():
        if s.code:
            subjects[s.code.lower()] = s
        subjects[s.name.lower()] = s

    academic_year = AcademicYear.query.filter_by(school_id=school_id, is_current=True).first()
    count = 0

    for row in rows:
        cls = classes.get(row.get('class_name', '').lower())
        sub = subjects.get(row.get('subject_code', '').lower())
        if not cls or not sub:
            continue

        syllabus = Syllabus(
            school_id=school_id,
            class_id=cls.id,
            subject_id=sub.id,
            academic_year_id=academic_year.id if academic_year else None,
            chapter_number=int(row['chapter_number']) if row.get('chapter_number', '').isdigit() else count + 1,
            chapter_name=row['chapter_name'],
            topics=row.get('topics') or None,
            learning_objectives=row.get('learning_objectives') or None,
            estimated_hours=float(row['estimated_hours']) if row.get('estimated_hours', '').replace('.', '').isdigit() else None,
            term=row.get('term', 'term1').lower(),
            status='not_started',
            completion_percentage=0,
        )
        db.session.add(syllabus)
        count += 1

    return count


def import_fee_structure(rows, school_id):
    """Import fee structure - auto creates fee categories"""
    classes = {c.name.lower(): c for c in Class.query.filter_by(school_id=school_id).all()}
    categories = {c.name.lower(): c for c in FeeCategory.query.filter_by(school_id=school_id).all()}
    academic_year = AcademicYear.query.filter_by(school_id=school_id, is_current=True).first()
    count = 0

    for row in rows:
        cls = classes.get(row.get('class_name', '').lower())
        if not cls:
            continue

        # Auto-create category if not exists
        cat_name = row.get('category_name', '').strip()
        if cat_name.lower() not in categories:
            cat = FeeCategory(school_id=school_id, name=cat_name)
            db.session.add(cat)
            db.session.flush()
            categories[cat_name.lower()] = cat

        cat = categories[cat_name.lower()]

        due_date = None
        if row.get('due_date'):
            try:
                due_date = datetime.strptime(row['due_date'], '%Y-%m-%d').date()
            except ValueError:
                pass

        fee = FeeStructure(
            school_id=school_id,
            academic_year_id=academic_year.id if academic_year else None,
            class_id=cls.id,
            fee_category_id=cat.id,
            amount=float(row['amount']),
            frequency=row.get('frequency', 'monthly').lower(),
            due_date=due_date,
            late_fee_amount=float(row['late_fee_amount']) if row.get('late_fee_amount', '').replace('.', '').isdigit() else 0,
            late_fee_type=row.get('late_fee_type', 'fixed').lower(),
            grace_period_days=int(row['grace_period_days']) if row.get('grace_period_days', '').isdigit() else 0,
            is_active=True,
        )
        db.session.add(fee)
        count += 1

    return count


def import_library_books(rows, school_id):
    """Import library books - auto creates categories"""
    categories = {c.name.lower(): c for c in LibraryCategory.query.filter_by(school_id=school_id).all()}
    count = 0

    for row in rows:
        # Auto-create category if provided and not exists
        cat_id = None
        if row.get('category_name'):
            cat_name = row['category_name'].strip()
            if cat_name.lower() not in categories:
                cat = LibraryCategory(school_id=school_id, name=cat_name)
                db.session.add(cat)
                db.session.flush()
                categories[cat_name.lower()] = cat
            cat_id = categories[cat_name.lower()].id

        copies = int(row['total_copies']) if row.get('total_copies', '').isdigit() else 1

        book = LibraryBook(
            school_id=school_id,
            category_id=cat_id,
            title=row['title'],
            author=row['author'],
            isbn=row.get('isbn') or None,
            publisher=row.get('publisher') or None,
            edition=row.get('edition') or None,
            language=row.get('language', 'English'),
            subject=row.get('subject') or None,
            publication_year=int(row['publication_year']) if row.get('publication_year', '').isdigit() else None,
            pages=int(row['pages']) if row.get('pages', '').isdigit() else None,
            total_copies=copies,
            available_copies=copies,
            rack_no=row.get('rack_no') or None,
            price=float(row['price']) if row.get('price', '').replace('.', '').isdigit() else None,
            is_active=True,
        )
        db.session.add(book)
        count += 1

    return count


def import_calendar_events(rows, school_id):
    """Import academic calendar events"""
    classes = {c.name.lower(): c for c in Class.query.filter_by(school_id=school_id).all()}
    event_colors = {'holiday': '#d32f2f', 'exam': '#ed6c02', 'ptm': '#9c27b0', 'event': '#1976d2',
                    'cultural': '#e91e63', 'sports': '#2e7d32', 'meeting': '#0288d1',
                    'deadline': '#f44336', 'vacation': '#ff9800', 'other': '#757575'}
    count = 0

    for row in rows:
        start = datetime.strptime(row['start_date'], '%Y-%m-%d').date()
        end = None
        if row.get('end_date'):
            try:
                end = datetime.strptime(row['end_date'], '%Y-%m-%d').date()
            except ValueError:
                pass

        class_id = None
        applies_to = row.get('applies_to', 'all').lower()
        if applies_to == 'specific_class' and row.get('class_name'):
            cls = classes.get(row['class_name'].lower())
            class_id = cls.id if cls else None

        etype = row.get('event_type', 'event').lower()
        event = AcademicCalendar(
            school_id=school_id,
            title=row['title'],
            description=row.get('description') or None,
            event_type=etype,
            start_date=start,
            end_date=end,
            is_holiday=row.get('is_holiday', 'false').lower() == 'true',
            applies_to=applies_to,
            class_id=class_id,
            color=event_colors.get(etype, '#1976d2'),
            notify_parents=False,
            created_by=g.current_user.id,
        )
        db.session.add(event)
        count += 1

    return count


# ============================================================
# IMPORT HISTORY / STATUS
# ============================================================

@imports_bp.route('/stats', methods=['GET'])
@school_required
def import_stats():
    """Get current data counts for the school - helps track migration progress"""
    stats = {
        'students': Student.query.filter_by(school_id=g.school_id, status='active').count(),
        'staff': Staff.query.filter_by(school_id=g.school_id, status='active').count(),
        'subjects': Subject.query.filter_by(school_id=g.school_id, is_active=True).count(),
        'classes': Class.query.filter_by(school_id=g.school_id).count(),
        'syllabus_chapters': Syllabus.query.filter_by(school_id=g.school_id).count(),
        'fee_categories': FeeCategory.query.filter_by(school_id=g.school_id).count(),
        'fee_structures': FeeStructure.query.filter_by(school_id=g.school_id).count(),
        'library_books': LibraryBook.query.filter_by(school_id=g.school_id).count(),
        'calendar_events': AcademicCalendar.query.filter_by(school_id=g.school_id).count(),
    }
    return success_response(stats)
